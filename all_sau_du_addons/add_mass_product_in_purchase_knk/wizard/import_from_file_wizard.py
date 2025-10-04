# -*- coding: utf-8 -*-
# Powered by Kanak Infosystems LLP.
# © 2020 Kanak Infosystems LLP. (<https://www.kanakinfosystems.com>).

import base64
import os
import tempfile
from openpyxl import load_workbook, Workbook
from io import BytesIO
import xlrd
import xlwt

from odoo import models, fields, _
from odoo.exceptions import ValidationError


class AddXls(models.TransientModel):
    _name = 'import.from.file.purchase'
    _description = "Import xls File"

    def _get_sample_file(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sample File"
        headers = [
            'Customer Product Code',
            'Customer Product Name',
            'Offered Description',
            'Quantity',
            'Unit Price',
            'Delivery Time',
        ]

        sheet.append(headers)

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return base64.encodebytes(stream.read())

    def _compute_sample_xls(self):
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet('Sample File')
        headers = [
            'Customer Product Code',
            'Customer Product Name',
            'Offered Description',
            'Quantity',
            'Unit Price',
            'Delivery Time',
        ]
        for col, header in enumerate(headers):
            sheet.write(0, col, header)

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        encoded = base64.encodebytes(stream.read())
        for record in self:
            record.download_sample_xls = encoded

    file = fields.Binary(string="Import File")
    filename = fields.Char()
    download_sample_xlsx = fields.Binary(
        string="Download XLSX Sample", default=_get_sample_file, readonly=True
    )
    download_sample_xls = fields.Binary(
        string="Download XLS Sample", compute="_compute_sample_xls", readonly=True
    )
    file_name_xlsx = fields.Char(default="Sample.xlsx")
    file_name_xls = fields.Char(default="Sample.xls")

    def _read_xls(self, file_data):
        order = self.env['purchase.order'].browse(self.env.context.get('active_ids'))

        try:
            decoded_data = base64.b64decode(file_data)
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xls")
            temp_file.write(decoded_data)
            temp_file.seek(0)

            workbook = xlrd.open_workbook(temp_file.name)
            sheet = workbook.sheet_by_index(0)
        except Exception as e:
            raise ValidationError(_("Invalid .xls file.\n\nTechnical error: %s") % str(e))

        values = []

        headers = [str(cell).strip().lower() for cell in sheet.row_values(0)]
        header_map = {
            'customer product code': 'sh_line_customer_code',
            'customer product name': 'sh_line_customer_product_name',
            'offered description': 'offered_description_id',
            # 'offered discription': 'offered_description_id',  # optional typo support
            'quantity': 'product_qty',
            'unit price': 'price_unit',
            'delivery time': 'delivery_time',
        }

        col_index_map = {}
        for idx, col in enumerate(headers):
            if col in header_map:
                col_index_map[header_map[col]] = idx

        for row_idx in range(1, sheet.nrows):
            row = sheet.row_values(row_idx)
            if not row or not any(row):
                continue

            def get_value(field):
                idx = col_index_map.get(field)
                return str(row[idx]).strip() if idx is not None and row[idx] is not None else ''

            customer_code = get_value('sh_line_customer_code')
            customer_name = get_value('sh_line_customer_product_name')
            offered_name = get_value('offered_description_id')
            quantity = float(get_value('product_qty') or 0)
            price = float(get_value('price_unit') or 0)
            delivery_time = int(float(get_value('delivery_time') or 0))

            # Try to match based on Offered Description first
            offered_product = self.env['product.product'].search([('name', '=', offered_name)], limit=1)

            # If not found, try matching by Customer Product Name
            if not offered_product and customer_name:
                offered_product = self.env['product.product'].search([('name', '=', customer_name)], limit=1)

            if not offered_product:
                line_vals = {
                    'product_uom_qty': quantity,
                    'price_unit': price,
                    'delivery_time': delivery_time,
                    'sh_line_customer_code': customer_code,
                    'sh_line_customer_product_name': customer_name,
                    'name': offered_name or customer_name or 'NA',
                }
            else:
                line_vals = {
                    'product_id': offered_product.id,
                    'offered_description_id': offered_product.id,
                    'product_uom_qty': quantity,
                    'price_unit': price,
                    'delivery_time': delivery_time,
                    'sh_line_customer_code': customer_code,
                    'sh_line_customer_product_name': customer_name,
                    'name': offered_product.display_name,
                }

            values.append((0, 0, line_vals))

        if values:
            order.write({'order_line': values})

    def _read_xlsx(self, file_data):
        order = self.env['purchase.order'].browse(self.env.context.get('active_ids'))

        try:
            decoded_data = base64.b64decode(file_data)
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            temp_file.write(decoded_data)
            temp_file.seek(0)

            workbook = load_workbook(filename=temp_file.name, data_only=True)
        except Exception as e:
            raise ValidationError(_("Invalid .xlsx file.\n\nTechnical error: %s") % str(e))

        sheet = workbook.active
        values = []

        headers = [str(cell.value).strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_map = {
            'customer product code': 'sh_line_customer_code',
            'customer product name': 'sh_line_customer_product_name',
            'offered description': 'offered_description_id',
            # 'offered discription': 'offered_description_id',  # optional typo support
            'quantity': 'product_qty',
            'unit price': 'price_unit',
            'delivery time': 'delivery_time',
        }

        col_index_map = {}
        for idx, col in enumerate(headers):
            if col in header_map:
                col_index_map[header_map[col]] = idx

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue

            def get_value(field):
                idx = col_index_map.get(field)
                return str(row[idx]).strip() if idx is not None and row[idx] is not None else ''

            customer_code = get_value('sh_line_customer_code')
            customer_name = get_value('sh_line_customer_product_name')
            offered_name = get_value('offered_description_id')
            quantity = float(get_value('product_qty') or 0)
            price = float(get_value('price_unit') or 0)
            delivery_time = int(float(get_value('delivery_time') or 0))

            offered_product = self.env['product.product'].search([('name', '=', offered_name)], limit=1)

            # If not found, try matching by Customer Product Name
            if not offered_product and customer_name:
                offered_product = self.env['product.product'].search([('name', '=', customer_name)], limit=1)

            if not offered_product:
                line_vals = {
                    'product_uom_qty': quantity,
                    'price_unit': price,
                    'delivery_time': delivery_time,
                    'sh_line_customer_code': customer_code,
                    'sh_line_customer_product_name': customer_name,
                    'name': offered_name or customer_name or 'NA',
                }
            else:
                line_vals = {
                    'product_id': offered_product.id,
                    'offered_description_id': offered_product.id,
                    'product_uom_qty': quantity,
                    'price_unit': price,
                    'delivery_time': delivery_time,
                    'sh_line_customer_code': customer_code,
                    'sh_line_customer_product_name': customer_name,
                    'name': offered_product.display_name,
                }

            values.append((0, 0, line_vals))

        if values:
            order.write({'order_line': values})

    def import_file(self):
        if not self.file:
            raise ValidationError(_("Please upload a file."))

        ext = os.path.splitext(self.filename)[1].lower()
        if ext not in ['.xlsx', '.xls']:
            raise ValidationError(_("Support Only Xlsx, Xls Formats."))

        if ext == '.xlsx':
            return self._read_xlsx(self.file)
        elif ext == '.xls':
            return self._read_xls(self.file)
