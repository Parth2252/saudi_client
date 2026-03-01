import openpyxl
from openpyxl.styles import Font, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# Headers
headers = ['S/N', 'TS CODE', 'PRODUCT DESCRIPTION', 'QTY', 'LOCATION']
ws.append(headers)

# Styling headers
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font

# Sample data
samples = [
    [1, 'TSHT06576', 'SHOVEL FLAT WITH WOODEN HANDLE', 10, 'A-05'],
    [2, 'TSCC37497', 'MAGNETIC WIRE 0.35MM', 5, 'A-01'],
    [3, 'TSHT24551', 'STEEL FLAT BAR', 20, 'B-01']
]

for row in samples:
    ws.append(row)

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[column].width = max_length + 2

wb.save('/home/pritesh/v18/oinc/saudi_client/stock_inventory_import/static/description/sample_inventory.xlsx')
