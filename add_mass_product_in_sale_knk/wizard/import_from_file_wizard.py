# -*- coding: utf-8 -*-
# Powered by Kanak Infosystems LLP.
# © 2020 Kanak Infosystems LLP. (<https://www.kanakinfosystems.com>).

import base64
import os
import tempfile
from openpyxl import load_workbook, Workbook
from io import BytesIO
import xlrd
import xlwt

from odoo import models, fields, _
from odoo.exceptions import ValidationError


class AddXls(models.TransientModel):
    _name = "import.from.file.sale"
    _description = "Import xls File"

    def _get_sample_file(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sample File"
        headers = [
            "Customer Product Code",
            "Customer Product Name",
            "Offered Description",
            "Quantity",
            "Unit Price",
            "Delivery Time",
            "Product Category",
            "UOM",
            "Vendor Name",
            "Vendor Price",
        ]

        sheet.append(headers)

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return base64.encodebytes(stream.read())

    def _compute_sample_xls(self):
        workbook = xlwt.Workbook()
        sheet = workbook.add_sheet("Sample File")
        headers = [
            "Customer Product Code",
            "Customer Product Name",
            "Offered Description",
            "Quantity",
            "Unit Price",
            "Delivery Time",
            "Product Category",
            "UOM",
            "Vendor Name",
            "Vendor Price",
        ]
        for col, header in enumerate(headers):
            sheet.write(0, col, header)

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        encoded = base64.encodebytes(stream.read())
        for record in self:
            record.download_sample_xls = encoded

    file = fields.Binary(string="Import File")
    filename = fields.Char()
    download_sample_xlsx = fields.Binary(
        string="Download XLSX Sample", default=_get_sample_file, readonly=True
    )
    download_sample_xls = fields.Binary(
        string="Download XLS Sample", compute="_compute_sample_xls", readonly=True
    )
    file_name_xlsx = fields.Char(default="Sample.xlsx")
    file_name_xls = fields.Char(default="Sample.xls")

    def _read_xls(self, file_data):
        order = self.env["sale.order"].browse(self.env.context.get("active_ids"))

        try:
            decoded_data = base64.b64decode(file_data)
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xls")
            temp_file.write(decoded_data)
            temp_file.seek(0)

            workbook = xlrd.open_workbook(temp_file.name)
            sheet = workbook.sheet_by_index(0)
        except Exception as e:
            raise ValidationError(
                _("Invalid .xls file.\n\nTechnical error: %s") % str(e)
            )

        values = []

        # Read header row (row 0) and normalize headers
        headers = [str(cell).strip().lower() for cell in sheet.row_values(0)]
        header_map = {
            "customer product code": "sh_line_customer_code",
            "customer product name": "sh_line_customer_product_name",
            "offered description": "offered_description_id",
            "quantity": "product_uom_qty",
            "unit price": "price_unit",
            "delivery time": "delivery_time",
            "product category": "product_category",
            "uom": "unit_of_measure",
            "vendor name": "vendor_name",
            "vendor price": "vendor_price",
            "n/a item": "na_item",  # NEW COLUMN
        }

        col_index_map = {}
        for idx, col in enumerate(headers):
            if col in header_map:
                col_index_map[header_map[col]] = idx

        # Parse each row starting from row 1
        for row_idx in range(1, sheet.nrows):
            row = sheet.row_values(row_idx)
            if not row or not any(row):
                continue

            def get_value(field):
                idx = col_index_map.get(field)
                return (
                    str(row[idx]).strip()
                    if idx is not None and row[idx] is not None
                    else ""
                )

            quantity = float(get_value("product_uom_qty") or 0)
            price = float(get_value("price_unit") or 0)
            delivery_time = int(float(get_value("delivery_time") or 0))

            # Customization start
            sh_product_customer_info_env = self.env["sh.product.customer.info"]
            product_env = self.env["product.product"]
            customer_product_name = get_value("sh_line_customer_product_name")
            customer_code = get_value("sh_line_customer_code")

            if customer_code:
                sh_product_customer_info_id = sh_product_customer_info_env.search(
                    [("product_code", "=", customer_code)], limit=1
                )
                product_id = sh_product_customer_info_id.product_id

                product_category = get_value("product_category")
                product_categ = self.env["product.category"]
                product_id = product_env.search(
                    [("name", "ilike", customer_product_name)], limit=1
                )
                product_category_id = product_categ.search(
                    [("name", "like", product_category)], limit=1
                )
                if not product_id:
                    product_id = product_env.create(
                        {"name": customer_product_name, "list_price": price}
                    )
                    if product_category_id:
                        product_id.write({"categ_id": product_category_id.id})

            if not customer_code:
                product_category = get_value("product_category")
                product_categ = self.env["product.category"]
                product_id = product_env.search(
                    [("name", "ilike", customer_product_name)], limit=1
                )
                product_category_id = product_categ.search(
                    [("name", "like", product_category)], limit=1
                )
                if not product_id:
                    product_id = product_env.create(
                        {"name": customer_product_name, "list_price": price}
                    )
                    if product_category_id:
                        product_id.write({"categ_id": product_category_id.id})
            # Customization end.

            # Try to match based on Offered Description first
            offered_name = get_value("offered_description_id").upper()
            offered_product_id = self.env["product.product"].search(
                [("name", "=", offered_name)], limit=1
            )

            product_category = get_value("product_category")
            product_categ = self.env["product.category"]
            product_category_id = product_categ.search(
                [("name", "like", product_category)], limit=1
            )

            if (
                not offered_product_id
                and offered_name not in (False, None, "", 0)
                and offered_name.lower()
                not in (
                    "na",
                    "not available",
                )
            ):
                offered_product_id = product_env.create(
                    {"name": offered_name, "list_price": price}
                )
                if product_category_id:
                    offered_product_id.write({"categ_id": product_category_id.id})

            uom = get_value("unit_of_measure")
            product_uom_id = False
            if uom:
                product_uom_id = self.env["uom.uom"].search(
                    [("name", "like", uom)], limit=1
                )

            if customer_code:
                sh_product_customer_info_id = sh_product_customer_info_env.search(
                    [
                        ("product_code", "=", customer_code),
                        ("name", "=", order.partner_id.id),
                    ],
                    limit=1,
                )
                # sh_pci = sh_product_customer_info_id.product_id.name
                if not sh_product_customer_info_id:
                    sh_product_id = product_id
                    # if offered_product_id:
                    #     sh_product_id = offered_product_id

                    sh_pci = sh_product_customer_info_env.create(
                        {
                            "name": order.partner_id.id,
                            "product_id": sh_product_id.id,
                            "product_tmpl_id": sh_product_id.product_tmpl_id.id,
                            "product_code": customer_code,
                            "product_name": customer_product_name,
                        }
                    )
                    customer_code = sh_pci.product_code

            # code for create product supplier info and link with product varient it self.
            product_supplierinfo_env = self.env["product.supplierinfo"]
            partner_env = self.env["res.partner"]
            vendor_name = get_value("vendor_name")
            vendor_price = float(get_value("vendor_price") or 0)

            vendor_id = partner_env.search([("name", "like", vendor_name)], limit=1)

            vendor_name_from_upload_excel = vendor_id
            vendor_price_from_upload_excel = vendor_price

            if offered_product_id and offered_name not in (False, None, "", 0):
                already_available_product_supplierinfo = (
                    product_supplierinfo_env.search(
                        [
                            ("partner_id", "=", vendor_id.id),
                            (
                                "product_tmpl_id",
                                "=",
                                offered_product_id.product_tmpl_id.id,
                            ),
                        ],
                        limit=1,
                    )
                )
            else:
                already_available_product_supplierinfo = (
                    product_supplierinfo_env.search(
                        [
                            ("partner_id", "=", vendor_id.id),
                            ("product_tmpl_id", "=", product_id.product_tmpl_id.id),
                        ],
                    )
                )

            if 1 <= delivery_time <= 7:
                exact_delivery_time = delivery_time - 1
            elif 8 <= delivery_time <= 14:
                exact_delivery_time = delivery_time - 2
            elif delivery_time >= 15:
                exact_delivery_time = delivery_time - 7
            else:
                exact_delivery_time = 1

            if not already_available_product_supplierinfo and not vendor_id and vendor_name:
                vendor_id = partner_env.create({"name": vendor_name, "is_vendor": True})

            if not already_available_product_supplierinfo and offered_product_id:
                product_supplierinfo_id = product_supplierinfo_env.create(
                    {
                        "partner_id": vendor_id.id,
                        "product_id": offered_product_id.id,
                        "min_qty": 1.0,
                        "price": vendor_price,
                        "delay": exact_delivery_time,
                    }
                )
            elif not already_available_product_supplierinfo and product_id:
                product_supplierinfo_id = product_supplierinfo_env.create(
                    {
                        "partner_id": vendor_id.id,
                        "product_id": product_id.id,
                        "min_qty": 1.0,
                        "price": vendor_price,
                        "delay": exact_delivery_time,
                    }
                )

            if vendor_name_from_upload_excel:
                product_vendor_id = vendor_name_from_upload_excel
            else:
                if already_available_product_supplierinfo:
                    product_vendor_id = (
                        already_available_product_supplierinfo.partner_id
                    )
                else:
                    product_vendor_id = product_supplierinfo_id.partner_id

            if vendor_price_from_upload_excel:
                vendor_price = vendor_price_from_upload_excel
            else:
                if already_available_product_supplierinfo:
                    vendor_price = already_available_product_supplierinfo[0].price or 0.0
                else:
                    vendor_price = 0.0

            # if related product is set then main product is set in product id field and related
            # product is set in offer description.
            if offered_product_id and offered_name.lower() not in (
                "na",
                "not available",
            ):
                line_vals = {
                    "product_id": product_id.id,
                    "offered_description_id": (offered_product_id.id),
                    "product_uom_qty": quantity,
                    "price_unit": 0.0,
                    "delivery_time": delivery_time,
                    "sh_line_customer_code": customer_code,
                    "sh_line_customer_product_name": (
                        offered_name if offered_product_id else customer_product_name
                    ),
                    "name": (
                        offered_product_id.name if offered_name else product_id.name
                    ),
                    "product_uom": product_uom_id.id if product_uom_id else False,
                    "product_vendor_id": product_vendor_id.id,
                    "vendor_price": vendor_price,
                }
            # if correct related product not set then execute below conditions.
            else:
                # if na or not available set in offered description then execute below condition.
                # unit price is zero because product is not available.
                if offered_name.lower() in ("na", "not available"):
                    line_vals = {
                        "product_id": product_id.id,
                        "product_uom_qty": quantity,
                        "price_unit": price,
                        "delivery_time": delivery_time,
                        "sh_line_customer_code": customer_code,
                        "sh_line_customer_product_name": (
                            offered_name
                            if offered_product_id
                            else customer_product_name
                        ),
                        "name": (product_id.name),
                        "is_not_available": True,
                        "product_uom": product_uom_id.id if product_uom_id else False,
                        "product_vendor_id": product_vendor_id.id,
                        "vendor_price": vendor_price,
                    }
                else:
                    # if no value is set in the offered description then execute this block.
                    line_vals = {
                        "product_id": product_id.id,
                        "product_uom_qty": quantity,
                        "price_unit": price,
                        "delivery_time": delivery_time,
                        "sh_line_customer_code": customer_code,
                        "sh_line_customer_product_name": (
                            offered_name
                            if offered_product_id
                            else customer_product_name
                        ),
                        "name": (product_id.name),
                        "product_uom": product_uom_id.id if product_uom_id else False,
                        "product_vendor_id": product_vendor_id.id,
                        "vendor_price": vendor_price,
                    }

            values.append((0, 0, line_vals))

        if values:
            order.write({"order_line": values})

    def _read_xlsx(self, file_data):
        order = self.env["sale.order"].browse(self.env.context.get("active_ids"))

        try:
            decoded_data = base64.b64decode(file_data)
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            temp_file.write(decoded_data)
            temp_file.seek(0)

            workbook = load_workbook(filename=temp_file.name, data_only=True)
        except Exception as e:
            raise ValidationError(
                _(
                    "Invalid file. Please upload a valid .xlsx file.\n\nTechnical error: %s"
                )
                % str(e)
            )

        sheet = workbook.active
        values = []

        # Read header row and map column names to indices
        headers = [
            str(cell.value).strip().lower()
            for cell in next(sheet.iter_rows(min_row=1, max_row=1))
        ]
        header_map = {
            "customer product code": "sh_line_customer_code",
            "customer product name": "sh_line_customer_product_name",
            "offered description": "offered_description_id",
            "quantity": "product_uom_qty",
            "unit price": "price_unit",
            "delivery time": "delivery_time",
            "product category": "product_category",
            "uom": "unit_of_measure",
            "vendor name": "vendor_name",
            "vendor price": "vendor_price",
            "n/a item": "na_item",  # NEW COLUMN
        }

        col_index_map = {}
        for idx, col in enumerate(headers):
            if col in header_map:
                col_index_map[header_map[col]] = idx

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue

            def get_value(field):
                idx = col_index_map.get(field)
                return (
                    str(row[idx]).strip()
                    if idx is not None and row[idx] is not None
                    else ""
                )

            quantity = float(get_value("product_uom_qty") or 0)
            price = float(get_value("price_unit") or 0)
            delivery_time = int(float(get_value("delivery_time") or 0))

            # Customization start
            sh_product_customer_info_env = self.env["sh.product.customer.info"]
            product_env = self.env["product.product"]
            customer_product_name = get_value("sh_line_customer_product_name")
            customer_code = get_value("sh_line_customer_code")

            if customer_code:
                sh_product_customer_info_id = sh_product_customer_info_env.search(
                    [("product_code", "=", customer_code)], limit=1
                )
                product_id = sh_product_customer_info_id.product_id

                product_category = get_value("product_category")
                product_categ = self.env["product.category"]
                product_id = product_env.search(
                    [("name", "ilike", customer_product_name)], limit=1
                )
                product_category_id = product_categ.search(
                    [("name", "like", product_category)], limit=1
                )
                if not product_id:
                    product_id = product_env.create(
                        {"name": customer_product_name, "list_price": price}
                    )
                    if product_category_id:
                        product_id.write({"categ_id": product_category_id.id})

            if not customer_code:
                product_category = get_value("product_category")
                product_categ = self.env["product.category"]
                product_id = product_env.search(
                    [("name", "ilike", customer_product_name)], limit=1
                )
                product_category_id = product_categ.search(
                    [("name", "like", product_category)], limit=1
                )
                if not product_id:
                    product_id = product_env.create(
                        {"name": customer_product_name, "list_price": price}
                    )
                    if product_category_id:
                        product_id.write({"categ_id": product_category_id.id})
            # Customization end.

            # Try to match based on Offered Description first
            offered_name = get_value("offered_description_id").upper()
            offered_product_id = self.env["product.product"].search(
                [("name", "=", offered_name)], limit=1
            )

            product_category = get_value("product_category")
            product_categ = self.env["product.category"]
            product_category_id = product_categ.search(
                [("name", "like", product_category)], limit=1
            )

            if (
                not offered_product_id
                and offered_name not in (False, None, "", 0)
                and offered_name.lower()
                not in (
                    "na",
                    "not available",
                )
            ):
                offered_product_id = product_env.create(
                    {"name": offered_name, "list_price": price}
                )
                if product_category_id:
                    offered_product_id.write({"categ_id": product_category_id.id})

            uom = get_value("unit_of_measure")
            product_uom_id = False
            if uom:
                product_uom_id = self.env["uom.uom"].search(
                    [("name", "like", uom)], limit=1
                )

            if customer_code:
                sh_product_customer_info_id = sh_product_customer_info_env.search(
                    [
                        ("product_code", "=", customer_code),
                        ("name", "=", order.partner_id.id),
                    ],
                    limit=1,
                )
                # sh_pci = sh_product_customer_info_id.product_id.name
                if not sh_product_customer_info_id:
                    sh_product_id = product_id
                    # if offered_product_id:
                    #     sh_product_id = offered_product_id

                    sh_pci = sh_product_customer_info_env.create(
                        {
                            "name": order.partner_id.id,
                            "product_id": sh_product_id.id,
                            "product_tmpl_id": sh_product_id.product_tmpl_id.id,
                            "product_code": customer_code,
                            "product_name": customer_product_name,
                        }
                    )
                    customer_code = sh_pci.product_code

            # code for create product supplier info and link with product varient it self.
            product_supplierinfo_env = self.env["product.supplierinfo"]
            partner_env = self.env["res.partner"]
            vendor_name = get_value("vendor_name")
            vendor_price = float(get_value("vendor_price") or 0)

            vendor_id = partner_env.search([("name", "like", vendor_name)], limit=1)

            vendor_name_from_upload_excel = vendor_id
            vendor_price_from_upload_excel = vendor_price

            if offered_product_id and offered_name not in (False, None, "", 0):
                already_available_product_supplierinfo = (
                    product_supplierinfo_env.search(
                        [
                            ("partner_id", "=", vendor_id.id),
                            (
                                "product_tmpl_id",
                                "=",
                                offered_product_id.product_tmpl_id.id,
                            ),
                        ],
                        limit=1,
                    )
                )
            else:
                already_available_product_supplierinfo = (
                    product_supplierinfo_env.search(
                        [
                            ("partner_id", "=", vendor_id.id),
                            ("product_tmpl_id", "=", product_id.product_tmpl_id.id),
                        ],
                    )
                )

            if 1 <= delivery_time <= 7:
                exact_delivery_time = delivery_time - 1
            elif 8 <= delivery_time <= 14:
                exact_delivery_time = delivery_time - 2
            elif delivery_time >= 15:
                exact_delivery_time = delivery_time - 7
            else:
                exact_delivery_time = 1

            if not already_available_product_supplierinfo and not vendor_id and vendor_name:
                vendor_id = partner_env.create({"name": vendor_name, "is_vendor": True})

            if not already_available_product_supplierinfo and offered_product_id:
                product_supplierinfo_id = product_supplierinfo_env.create(
                    {
                        "partner_id": vendor_id.id,
                        "product_id": offered_product_id.id,
                        "min_qty": 1.0,
                        "price": vendor_price,
                        "delay": exact_delivery_time,
                    }
                )
            elif not already_available_product_supplierinfo and offered_product_id:
                product_supplierinfo_id = product_supplierinfo_env.create(
                    {
                        "partner_id": vendor_id.id,
                        "product_id": product_id.id,
                        "min_qty": 1.0,
                        "price": vendor_price,
                        "delay": exact_delivery_time,
                    }
                )

            if vendor_name_from_upload_excel:
                product_vendor_id = vendor_name_from_upload_excel
            else:
                if already_available_product_supplierinfo:
                    product_vendor_id = (
                        already_available_product_supplierinfo.partner_id
                    )
                else:
                    product_vendor_id = product_supplierinfo_id.partner_id

            if vendor_price_from_upload_excel:
                vendor_price = vendor_price_from_upload_excel
            else:
                if already_available_product_supplierinfo:
                    vendor_price = already_available_product_supplierinfo[0].price or 0.0
                else:
                    vendor_price = 0.0

            # if related product is set then main product is set in product id field and related
            # product is set in offer description.
            if offered_product_id and offered_name.lower() not in (
                "na",
                "not available",
            ):
                line_vals = {
                    "product_id": product_id.id,
                    "offered_description_id": (offered_product_id.id),
                    "product_uom_qty": quantity,
                    "price_unit": price,
                    "delivery_time": delivery_time,
                    "sh_line_customer_code": customer_code,
                    "sh_line_customer_product_name": (
                        offered_name if offered_product_id else customer_product_name
                    ),
                    "name": (
                        offered_product_id.name if offered_name else product_id.name
                    ),
                    "product_uom": product_uom_id.id if product_uom_id else False,
                    "product_vendor_id": product_vendor_id.id,
                    "vendor_price": vendor_price,
                }
            # if correct related product not set then execute below conditions.
            else:
                # if na or not available set in offered description then execute below condition.
                # unit price is zero because product is not available.
                if offered_name.lower() in ("na", "not available"):
                    line_vals = {
                        "product_id": product_id.id,
                        "product_uom_qty": quantity,
                        "price_unit": price,
                        "delivery_time": delivery_time,
                        "sh_line_customer_code": customer_code,
                        "sh_line_customer_product_name": (
                            offered_name
                            if offered_product_id
                            else customer_product_name
                        ),
                        "name": (product_id.name),
                        "is_not_available": True,
                        "product_uom": product_uom_id.id if product_uom_id else False,
                        "product_vendor_id": product_vendor_id.id,
                        "vendor_price": vendor_price,
                    }
                else:
                    # if no value is set in the offered description then execute this block.
                    line_vals = {
                        "product_id": product_id.id,
                        "product_uom_qty": quantity,
                        "price_unit": price,
                        "delivery_time": delivery_time,
                        "sh_line_customer_code": customer_code,
                        "sh_line_customer_product_name": (
                            offered_name
                            if offered_product_id
                            else customer_product_name
                        ),
                        "name": (product_id.name),
                        "product_uom": product_uom_id.id if product_uom_id else False,
                        "product_vendor_id": product_vendor_id.id,
                        "vendor_price": vendor_price,
                    }

            values.append((0, 0, line_vals))

        if values:
            order.write({"order_line": values})

    def import_file(self):
        if not self.file:
            raise ValidationError(_("Please upload a file."))

        ext = os.path.splitext(self.filename)[1].lower()
        if ext not in [".xlsx", ".xls"]:
            raise ValidationError(_("Support Only Xlsx, Xls Formats."))

        if ext == ".xlsx":
            return self._read_xlsx(self.file)
        elif ext == ".xls":
            return self._read_xls(self.file)
